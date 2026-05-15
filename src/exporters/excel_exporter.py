# src/exporters/excel_exporter.py
"""
Excel Exporter module.
Serializes structural DataFrames to an Excel workbook via independent sheets.
"""

import pandas as pd
import logging
import os
from typing import Any
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, config: dict[str, Any] = None):
        self.config = config or {}

    def _auto_fit_columns(self, ws) -> None:
        """Dynamically adjusts column widths to prevent structural text truncation."""
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    def _apply_autofilter(self, ws) -> None:
        """Applies dynamic header filters to tabular data matrices."""
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

    def _apply_risk_color_coding(self, ws) -> None:
        """Injects conditional fill patterns for Supplier Risk identification."""
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        risk_col_idx = None
        for col in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=1, column=col).value or "").upper()
            if 'RISK' in header_val:
                risk_col_idx = col
                break

        if not risk_col_idx:
            return

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=risk_col_idx)
            val = str(cell.value).upper() if cell.value else ""
            if 'HIGH' in val:
                cell.fill = red_fill
            elif 'MEDIUM' in val:
                cell.fill = yellow_fill
            elif 'LOW' in val:
                cell.fill = green_fill

    def export(self, results: dict[str, Any], output_path: str) -> None:
        try:
            os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

                if 'summary_metadata' in results:
                    meta_df = pd.DataFrame(list(results['summary_metadata'].items()), columns=['Metric', 'Value'])
                    meta_df.to_excel(writer, sheet_name='Summary', index=False)
                    self._auto_fit_columns(writer.sheets['Summary'])
                    # No autofilter for summary sheet

                target_sheets = [
                    ('abc_classification', 'ABC Classification'),
                    ('monthly_forecast', 'Monthly Forecast'),
                    ('quarterly_forecast', 'Quarterly Forecast'),
                    ('exceptions', 'Delivery Exceptions'),
                    ('supplier_risk_report', 'Supplier Risk Report'),
                    ('monthly_breakdown', 'Monthly Breakdown'),
                    ('lead_times', 'Lead Times')
                ]

                for key, sheet_name in target_sheets:
                    if key in results and not results[key].empty:
                        results[key].to_excel(writer, sheet_name=sheet_name, index=False)
                        ws = writer.sheets[sheet_name]
                        self._auto_fit_columns(ws)
                        self._apply_autofilter(ws)

                        if key == 'supplier_risk_report':
                            self._apply_risk_color_coding(ws)

                raw_sheets = {
                    'Raw GRN': 'raw_grn',
                    'Raw POV': 'raw_pov',
                    'Raw PV': 'raw_pv'
                }
                for sheet_name, key in raw_sheets.items():
                    if key in results and not results[key].empty:
                        results[key].to_excel(writer, sheet_name=sheet_name, index=False)
                        ws = writer.sheets[sheet_name]
                        self._auto_fit_columns(ws)
                        self._apply_autofilter(ws)

            logger.info("Excel artifact successfully generated at: %s", output_path)

        except Exception as e:
            logger.error("Failed to execute Excel serialization: %s", str(e))
            raise